#!/usr/bin/env python3
"""Parse the UC Berkeley deep-dive sources into a tidy CSV.

Primary: outcomes data/berkeley_disaggregated_grad_rates_final/normalized_rates.csv
 — a VizQL extraction of Berkeley OPAP's "Disaggregated Grad Rates" dashboard
 (calviz.berkeley.edu): freshman & transfer entrants 2010-2023, full-precision
 graduation rates WITH numerators and denominators, by Overall, detailed
 race/ethnicity (incl. Asian and Latinx subgroups), gender, first-generation,
 and EOP eligibility. Year-by-Year mode only (the 5-yr moving average is a
 derivable smoothing). Window gates: FR 6-yr <= 2019, 4-yr <= 2021; TR 2-yr
 <= 2023, 4-yr <= 2021 (vs the 2025 data vintage).

Secondary: outcomes data/UCB Students - Graduation & Retention Rates.xlsx
 ("Our Berkeley" export): one-year retention counts by ethnicity x gender x
 residency, and by ENTRY COLLEGE. Rates = Retained / (Retained + Not Retained),
 aggregated over residency. Provides dimensions calviz lacks: entry college and
 the ethnicity x gender cross (coarse ethnicity).

Output: data/trends/ucb_rates.csv
  level, dimension, subgroup, cohort, measure, rate_pct, numer, denom
"""
import argparse, csv, json, os
from collections import defaultdict, Counter

MET = {'4-year rate': 'grad4', '6-year rate': 'grad6', '2-year rate': 'grad2'}
GATE = {('FR', 'grad6'): 2019, ('FR', 'grad4'): 2021,
        ('TR', 'grad2'): 2023, ('TR', 'grad4'): 2021, ('TR', 'grad6'): 2019}
DIM = {'Overall': 'overall', 'Race/Ethnicity': 'eth_detail', 'Gender': 'gender',
       'First Generation': 'firstgen', 'EOP Eligibility': 'eop'}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--corpus', default=r'C:\Users\harth\svetlana\outcomes data')
    ap.add_argument('--out', default=r'C:\Users\harth\svetlana\Svetlana\uc-merit-admissions\data\trends')
    a = ap.parse_args()
    out_rows = []
    qa = Counter()

    # ---- calviz disaggregated grad rates ----
    src = os.path.join(a.corpus, 'berkeley_disaggregated_grad_rates_final', 'normalized_rates.csv')
    best = {}
    for r in csv.DictReader(open(src, encoding='utf-8')):
        if r['measure_mode'] != 'Year by Year':
            continue
        meas = MET.get(r['metric'])
        if meas is None:
            qa['calviz_skipped_metric'] += 1
            continue
        if r['rate'] in ('', None) or r['denominator'] in ('', None):
            qa['calviz_no_rate'] += 1
            continue
        level = {'Freshman': 'FR', 'Transfer': 'TR'}.get(r['student_level'])
        if level is None:  # Overview rows: entry type lives in the raw mark
            try:
                et = json.loads(r['raw_row_json']).get('Entry Type', '')
            except Exception:
                et = ''
            level = {'Freshman Entrant': 'FR', 'Transfer Entrant': 'TR'}.get(et)
            if level is None:
                qa['calviz_no_level'] += 1
                continue
        dim = DIM.get(r['demographic_dimension'])
        if dim is None:
            qa['calviz_skipped_dim'] += 1
            continue
        sub = (r['display_category'] or 'Overall').strip()
        if dim == 'overall' or sub in ('4-Year Rate', '6-Year Rate', '2-Year Rate', ''):
            dim, sub = 'overall', 'Overall'
        cohort = int(r['cohort'])
        if cohort > GATE.get((level, meas), 2023):
            qa['calviz_gated_window'] += 1
            continue
        rate = float(r['rate'])
        den = int(round(float(r['denominator'])))
        num = int(round(float(r['numerator'] or 0)))
        k = (level, dim, sub, cohort, meas)
        if k in best and abs(best[k][0] - rate) > 1e-9:
            qa['calviz_conflict'] += 1
        best[k] = (rate, num, den)
    for (level, dim, sub, cohort, meas), (rate, num, den) in best.items():
        out_rows.append([level, dim, sub, cohort, meas, round(rate * 100, 2), num, den])
    qa['calviz_rows'] = len(best)

    # ---- Our Berkeley 1-yr retention ----
    import openpyxl
    wb = openpyxl.load_workbook(
        os.path.join(a.corpus, 'UCB Students - Graduation & Retention Rates.xlsx'), read_only=True)

    def cohort_of(ay):  # '2014-15' -> 2014
        return int(str(ay).split('-')[0])

    def lv(et):
        return {'Freshman Entrant': 'FR', 'Transfer Entrant': 'TR'}.get(str(et).strip())

    # sheet: eth x gender x residency
    agg = defaultdict(lambda: [0, 0])   # key -> [retained, total]
    ws = wb['1-Yr Retention Ethnicity-Gender']
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(x) for x in row]
            continue
        d = dict(zip(hdr, row))
        level = lv(d.get('Entry Type'))
        if level is None or d.get('Headcount') in (None, ''):
            continue
        coh = cohort_of(d['Academic Yr'])
        n = int(d['Headcount'])
        ret = 1 if str(d.get('Status')).strip() == 'Retained' else 0
        eth = str(d['Ethnicity']).strip()
        gen = str(d['Gender']).strip()
        for key in ((level, 'ret_eth', eth, coh), (level, 'ret_gender', gen, coh),
                    (level, 'ret_ethgen', eth + ' × ' + gen, coh),
                    (level, 'ret_overall', 'Overall', coh)):
            agg[key][0] += n * ret
            agg[key][1] += n
    # sheet: entry college
    ws = wb['1-Yr Retention Entry Clg']
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(x) for x in row]
            continue
        d = dict(zip(hdr, row))
        level = lv(d.get('Entry Type'))
        if level is None or d.get('Headcount') in (None, ''):
            continue
        coh = cohort_of(d['Academic Yr'])
        n = int(d['Headcount'])
        ret = 1 if str(d.get('Status')).strip() == 'Retained' else 0
        clg = str(d['Entry College']).strip()
        agg[(level, 'ret_college', clg, coh)][0] += n * ret
        agg[(level, 'ret_college', clg, coh)][1] += n
    MIN_N = 20   # hide tiny cells (source is unsuppressed counts)
    for (level, dim, sub, coh), (num, den) in agg.items():
        if den < MIN_N:
            qa['ourb_small_cell'] += 1
            continue
        out_rows.append([level, dim, sub, coh, 'ret1', round(100.0 * num / den, 2), num, den])
    qa['ourb_rows'] = sum(1 for r in out_rows if r[4] == 'ret1')

    out_rows.sort(key=lambda r: (r[0], r[1], r[2], r[4], r[3]))
    os.makedirs(a.out, exist_ok=True)
    with open(os.path.join(a.out, 'ucb_rates.csv'), 'w', newline='') as fh:
        w = csv.writer(fh)
        w.writerow(['level', 'dimension', 'subgroup', 'cohort', 'measure', 'rate_pct', 'numer', 'denom'])
        w.writerows(out_rows)
    meta = {'rows': len(out_rows), 'qa': dict(qa),
            'dims': {d: sorted({r[2] for r in out_rows if r[1] == d})
                     for d in sorted({r[1] for r in out_rows})},
            'cohorts': [min(r[3] for r in out_rows), max(r[3] for r in out_rows)]}
    with open(os.path.join(a.out, 'ucb_qa.json'), 'w') as fh:
        json.dump(meta, fh, indent=1, ensure_ascii=False)
    print(json.dumps({'rows': meta['rows'], 'qa': meta['qa'], 'cohorts': meta['cohorts']}))
    for d, subs in meta['dims'].items():
        print(' ', d, len(subs), subs[:8])

if __name__ == '__main__':
    main()
